import csv
import io
import json
from decimal import Decimal
from datetime import date, datetime

from app.repositories import lancamentos_repository, contas_repository, config_repository


def _serialize(value):
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return value


def _lancamentos_rows(user_id: int) -> list[dict]:
    rows = lancamentos_repository.get_lancamentos_by_user(user_id)
    return [
        {
            "id": r["id"],
            "tipo": r["tipo"],
            "descricao": r["descricao"],
            "valor": float(r["valor"] or 0),
            "data_vencimento": r["data_vencimento"].isoformat() if r.get("data_vencimento") else "",
            "data_pagamento": r["data_pagamento"].isoformat() if r.get("data_pagamento") else "",
            "status": r.get("status", ""),
            "categoria": r.get("categoria_nome", ""),
            "subcategoria": r.get("subcategoria_nome", ""),
            "conta": r.get("conta_nome", ""),
            "cartao": r.get("cartao_nome", ""),
            "recorrente": r.get("recorrente", False),
            "observacao": r.get("observacao", ""),
        }
        for r in rows
    ]


def _contas_rows(user_id: int) -> list[dict]:
    rows = contas_repository.list_contas(user_id)
    return [
        {
            "id": r["id"],
            "nome": r["nome"],
            "instituicao": r.get("instituicao", ""),
            "saldo_inicial": float(r.get("saldo_inicial") or 0),
        }
        for r in rows
    ]


def _categorias_rows(user_id: int) -> list[dict]:
    result = []
    for tipo in ("despesa", "receita", "conta"):
        for cat in config_repository.get_categorias(user_id, tipo):
            result.append({
                "tipo": tipo,
                "nome": cat["nome"],
                "icone": cat.get("icone", ""),
            })
    return result


_EXPORTERS = {
    "lancamentos": _lancamentos_rows,
    "contas": _contas_rows,
    "categorias": _categorias_rows,
}


def export_csv(user_id: int, tipo: str) -> tuple[bytes, str]:
    getter = _EXPORTERS.get(tipo)
    if not getter:
        raise ValueError("Tipo inválido")
    rows = getter(user_id)
    if not rows:
        rows = [{}]
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(rows[0].keys()) if rows else [])
    writer.writeheader()
    writer.writerows(rows)
    filename = f"fluxara_{tipo}_{date.today().isoformat()}.csv"
    return buf.getvalue().encode("utf-8-sig"), filename


def export_json(user_id: int, tipo: str) -> tuple[bytes, str]:
    getter = _EXPORTERS.get(tipo)
    if not getter:
        raise ValueError("Tipo inválido")
    rows = getter(user_id)
    content = json.dumps(rows, ensure_ascii=False, indent=2)
    filename = f"fluxara_{tipo}_{date.today().isoformat()}.json"
    return content.encode("utf-8"), filename


def export_xlsx(user_id: int, tipo: str) -> tuple[bytes, str]:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise RuntimeError("openpyxl não instalado")

    getter = _EXPORTERS.get(tipo)
    if not getter:
        raise ValueError("Tipo inválido")
    rows = getter(user_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()

    if not rows:
        filename = f"fluxara_{tipo}_{date.today().isoformat()}.xlsx"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), filename

    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"fluxara_{tipo}_{date.today().isoformat()}.xlsx"
    return buf.getvalue(), filename
